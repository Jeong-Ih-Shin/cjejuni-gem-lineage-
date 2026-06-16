"""
Step 04 — Figure 2: In vitro phenotype lineage signal (24h fc).

Inputs:
    data/metadata.xlsx     (Long_format + Cluster_assignments)

Outputs:
    output/figures/Figure2_invitro_profiles.{png,svg}
"""
from pathlib import Path
import sys
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.preprocessing import StandardScaler
from scipy.spatial.distance import pdist, squareform, cdist
from scipy.stats import mannwhitneyu
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.colors import LinearSegmentedColormap

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import DATA_DIR, NPG, SEM, CHEMICALS, apply_style, save_fig

apply_style()
HIGHLIGHT = NPG['green']   # teal — distinguishes Fig 2 from Fig 1 (red)


def main():
    wb = load_workbook(DATA_DIR / 'metadata.xlsx', read_only=True)
    ws = wb['Cluster_assignments']
    rows = list(ws.iter_rows(values_only=True))
    v4 = pd.DataFrame(rows[1:], columns=rows[0]).dropna(subset=['sample'])
    v4 = v4[~v4['sample'].str.contains('KCTC|ATCC', case=False, na=False)]

    ws2 = wb['Long_format']
    rows2 = list(ws2.iter_rows(values_only=True))
    phen = pd.DataFrame(rows2[1:], columns=rows2[0])
    phen['fc_trimmed'] = pd.to_numeric(phen['fc_trimmed'], errors='coerce')
    phen['timepoint']  = pd.to_numeric(phen['timepoint'],  errors='coerce')
    phen = phen.dropna(subset=['fc_trimmed', 'timepoint'])
    phen = phen[~phen['sample'].str.contains('KCTC|ATCC', case=False, na=False)]

    # 24h fc per (strain, chemical)
    records = []
    for sample, sub in phen.groupby('sample'):
        for chem in CHEMICALS:
            cs = sub[(sub.chemical == chem) & (sub.timepoint == 24)]
            if len(cs) == 0:
                continue
            records.append({'sample': sample, 'chemical': chem, 'val': cs['fc_trimmed'].iloc[0]})
    wide = pd.DataFrame(records).pivot(index='sample', columns='chemical', values='val').dropna()
    wide['CC'] = wide.index.map(dict(zip(v4['sample'], v4['CC'])))
    print(f'Strains with complete 24h fc: {len(wide)}')

    # z-score per chemical (cohort-wide)
    z = StandardScaler().fit_transform(wide[CHEMICALS].values)
    z_df = pd.DataFrame(z, index=wide.index, columns=CHEMICALS)
    z_df['CC'] = wide['CC']

    # Pairwise Euclidean
    dist_arr = pdist(z_df[CHEMICALS].values, metric='euclidean')
    dist_mat = pd.DataFrame(squareform(dist_arr), index=z_df.index, columns=z_df.index)
    samples = z_df.index.tolist()
    pairs = []
    for i, s1 in enumerate(samples):
        for j, s2 in enumerate(samples):
            if i >= j:
                continue
            c1, c2 = z_df.loc[s1, 'CC'], z_df.loc[s2, 'CC']
            if pd.isna(c1) or pd.isna(c2):
                continue
            pairs.append({'same_cc': c1 == c2, 'dist': dist_mat.iloc[i, j]})
    pairs_df = pd.DataFrame(pairs)
    same = pairs_df[pairs_df.same_cc]['dist'].values
    diff = pairs_df[~pairs_df.same_cc]['dist'].values
    _, p = mannwhitneyu(same, diff, alternative='less')
    med_same, med_diff = np.median(same), np.median(diff)
    print(f'Same-CC: n={len(same)}, median={med_same:.2f}')
    print(f'Diff-CC: n={len(diff)}, median={med_diff:.2f}')
    print(f'Mann-Whitney p = {p:.2e}')

    cc_counts = z_df['CC'].value_counts()
    coherent = ['ST-21 complex', 'ST-443 complex', 'ST-658 complex',
                'ST-607 complex', 'ST-45 complex', 'ST-354 complex']
    non_coherent = ['ST-464 complex', 'ST-48 complex', 'ST-5229 complex']
    ordered = coherent + non_coherent

    fig, axes = plt.subplots(1, 2, figsize=(6.69, 3.5),
                              gridspec_kw={'width_ratios': [1.3, 1.0]})
    plt.subplots_adjust(top=0.86, bottom=0.18, left=0.05, right=0.97, wspace=0.5)


    # B — heatmap
    ax = axes[0]
    ax.grid(False)
    heatmap_data, cc_labels = [], []
    for c in ordered:
        s = z_df[z_df.CC == c]
        if len(s) == 0:
            continue
        heatmap_data.append(s[CHEMICALS].mean().values)
        cc_labels.append(f'{c.replace(" complex","")} (n={len(s)})')
    cmap = LinearSegmentedColormap.from_list('npg_div', [SEM['down'], '#FFFFFF', SEM['up']], N=256)
    sns.heatmap(np.array(heatmap_data), annot=True, fmt='+.2f', cmap=cmap,
                center=0, vmin=-1.5, vmax=1.5,
                xticklabels=CHEMICALS, yticklabels=cc_labels, ax=ax,
                cbar_kws={'label': 'Mean z-score'}, linewidths=1.5, linecolor='white',
                annot_kws={'color': SEM['text'], 'fontsize': 7, 'fontweight': 'bold'})
    n_coh = sum(1 for c in coherent if c in cc_counts.index)
    ax.set_title('A', loc='left', x=-0.30, fontsize=11, fontweight='bold')
    ax.tick_params(axis='x', rotation=30)

    # C — per-CC bars
    ax = axes[1]
    per_cc = []
    for c in cc_counts[cc_counts >= 2].index:
        if pd.isna(c) or c is None:
            continue
        in_cc = z_df[z_df.CC == c][CHEMICALS].values
        out_cc = z_df[z_df.CC != c][CHEMICALS].values
        if len(in_cc) < 2:
            continue
        within = pdist(in_cc).mean()
        between = cdist(in_cc, out_cc).mean()
        per_cc.append({'cc': f'{c.replace(" complex","")} (n={cc_counts[c]})',
                       'within': within, 'between': between, 'n': cc_counts[c]})
    per_cc_df = pd.DataFrame(per_cc).sort_values('n', ascending=True)
    y_pos = np.arange(len(per_cc_df))
    ax.barh(y_pos - 0.20, per_cc_df['within'], 0.38, color=HIGHLIGHT,
            label='Within-CC mean', edgecolor='white', linewidth=1.2)
    ax.barh(y_pos + 0.20, per_cc_df['between'], 0.38, color=SEM['neutral'],
            label='Between-CC mean', edgecolor='white', linewidth=1.2)
    ax.set_yticks(y_pos)
    ax.set_yticklabels(per_cc_df['cc'])
    ax.set_xlabel('Mean Euclidean distance')
    ax.set_title('B', loc='left', x=-0.30, fontsize=11, fontweight='bold')
    ax.legend(frameon=False, loc='lower center', bbox_to_anchor=(0.5, 1.02),
              ncol=2, fontsize=7)
    ax.grid(axis='x', alpha=0.5, color=SEM['grid'], linewidth=0.5)
    ax.set_xlim(0, max(per_cc_df['between'].max(), per_cc_df['within'].max()) * 1.10)

    save_fig(fig, 'Figure2_invitro_profiles')
    plt.close(fig)
    print('Done.')


if __name__ == '__main__':
    main()
