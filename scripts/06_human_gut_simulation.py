"""
Step 07: In silico FBA-based growth simulation in human gut media.

Inputs:
    gems/*.xml             (one SBML per strain)
    data/metadata.xlsx     (Cluster_assignments sheet)

Outputs:
    intermediate/human_gut_simulation_results.csv

Method:
    Two diet-mimicking media (adapted from Magnusdottir et al. 2017 Nat Biotechnol):
      - Western diet     : high in simple sugars, fats, amino acids; lower fiber
      - High-fiber diet  : enriched in fermentation products (SCFAs, lactate, formate)
    For each strain x each medium, close all exchanges then open diet uptakes.
    Microaerobic O2 uptake capped at 5 mmol/gDW/h.
    Objective: maximize the biomass-forming reaction (yielding ModelSEED
    biomass pseudo-metabolite cpd11416_c0).
    Record the maximum theoretical growth rate mu (1/h) for each (strain, medium).
"""

from pathlib import Path

import numpy as np
import pandas as pd
import cobra
from openpyxl import load_workbook

# ============================================================
# Paths
# ============================================================
REPO_ROOT        = Path(__file__).resolve().parent.parent
GEMS_DIR         = REPO_ROOT / 'gems'
DATA_DIR         = REPO_ROOT / 'data'
INTERMEDIATE_DIR = REPO_ROOT / 'intermediate'
INTERMEDIATE_DIR.mkdir(exist_ok=True)

METADATA_XLSX = DATA_DIR / 'metadata.xlsx'
OUTPUT_CSV    = INTERMEDIATE_DIR / 'human_gut_simulation_results.csv'


# ============================================================
# Medium definitions (Magnusdottir et al. 2017, adapted)
# ============================================================

UNLIMITED_IONS = {
    'cpd00001': 1000, 'cpd00009': 1000, 'cpd00011': 1000, 'cpd00013': 1000,
    'cpd00030': 1000, 'cpd00034': 1000, 'cpd00048': 1000, 'cpd00058': 1000,
    'cpd00063': 1000, 'cpd00067': 1000, 'cpd00099': 1000, 'cpd00149': 1000,
    'cpd00205': 1000, 'cpd00254': 1000, 'cpd00971': 1000,
    'cpd10515': 1000, 'cpd10516': 1000,
}


WESTERN_DIET = {
    **UNLIMITED_IONS,
    # Simple carbohydrates (rich)
    'cpd00027': 5.0,   # Glucose
    'cpd00082': 1.0,   # Fructose
    'cpd00108': 1.0,   # Galactose
    'cpd00179': 1.0,   # Maltose
    'cpd00208': 1.0,   # Lactose
    'cpd00076': 1.0,   # Sucrose
    'cpd00138': 0.5,   # Mannose
    'cpd00159': 1.0,   # Lactate
    'cpd00609': 0.5,   # Gluconate
    'cpd00047': 1.0,   # Formate
    # SCFA (moderate)
    'cpd00029': 5.0,   # Acetate
    'cpd00141': 1.0,   # Propionate
    'cpd00211': 1.0,   # Butyrate
    # 20 amino acids
    'cpd00033': 0.5, 'cpd00035': 0.5, 'cpd00039': 0.5, 'cpd00041': 0.5,
    'cpd00051': 0.5, 'cpd00053': 0.5, 'cpd00054': 0.5, 'cpd00060': 0.5,
    'cpd00065': 0.5, 'cpd00066': 0.5, 'cpd00069': 0.5, 'cpd00084': 0.5,
    'cpd00107': 0.5, 'cpd00119': 0.5, 'cpd00129': 0.5, 'cpd00132': 0.5,
    'cpd00156': 0.5, 'cpd00161': 0.5, 'cpd00322': 0.5, 'cpd00023': 0.5,
    # B-vitamins
    'cpd00644': 0.05, 'cpd00104': 0.01, 'cpd00220': 0.01, 'cpd00305': 0.01,
    'cpd00263': 0.01, 'cpd00393': 0.01, 'cpd00078': 0.01, 'cpd00133': 0.01,
    'cpd00028': 0.01,
    # Lipid / osmolyte
    'cpd00100': 0.5, 'cpd00136': 0.1, 'cpd00154': 0.05,
    # Mucin sugars (modest)
    'cpd00131': 0.1, 'cpd03847': 0.05, 'cpd00751': 0.05,
    # Sulfur
    'cpd00081': 0.1, 'cpd00268': 0.05,
    # O2 (microaerobic)
    'cpd00007': 5.0,
}


HIGH_FIBER_DIET = {
    **UNLIMITED_IONS,
    # Less simple carbs, more complex carbs/SCFAs
    'cpd00027': 1.0,   # Glucose (reduced)
    'cpd00082': 0.5,   # Fructose
    'cpd00108': 0.5,   # Galactose
    'cpd00179': 0.5,   # Maltose
    'cpd00208': 0.5,   # Lactose
    'cpd00076': 0.5,   # Sucrose
    'cpd00138': 0.5,   # Mannose
    'cpd00159': 1.5,   # Lactate (elevated from fermentation)
    'cpd00609': 0.5,   # Gluconate
    'cpd00047': 2.0,   # Formate (elevated from fermentation)
    'cpd00072': 1.0,   # Cellobiose (fiber breakdown)
    # SCFA (high — fiber fermentation)
    'cpd00029': 10.0,  # Acetate (high)
    'cpd00141': 3.0,   # Propionate (elevated)
    'cpd00211': 3.0,   # Butyrate (elevated)
    # 20 amino acids (reduced — less protein in high-fiber diet)
    'cpd00033': 0.3, 'cpd00035': 0.3, 'cpd00039': 0.3, 'cpd00041': 0.3,
    'cpd00051': 0.3, 'cpd00053': 0.3, 'cpd00054': 0.3, 'cpd00060': 0.3,
    'cpd00065': 0.3, 'cpd00066': 0.3, 'cpd00069': 0.3, 'cpd00084': 0.3,
    'cpd00107': 0.3, 'cpd00119': 0.3, 'cpd00129': 0.3, 'cpd00132': 0.3,
    'cpd00156': 0.3, 'cpd00161': 0.3, 'cpd00322': 0.3, 'cpd00023': 0.3,
    # B-vitamins
    'cpd00644': 0.05, 'cpd00104': 0.01, 'cpd00220': 0.01, 'cpd00305': 0.01,
    'cpd00263': 0.01, 'cpd00393': 0.01, 'cpd00078': 0.01, 'cpd00133': 0.01,
    'cpd00028': 0.01,
    # Lipid / osmolyte
    'cpd00100': 0.5, 'cpd00136': 0.1, 'cpd00154': 0.1,  # more betaine (plant-derived)
    # Mucin sugars
    'cpd00131': 0.1, 'cpd03847': 0.05, 'cpd00751': 0.05,
    # Sulfur
    'cpd00081': 0.1, 'cpd00268': 0.05,
    # O2 (microaerobic)
    'cpd00007': 5.0,
}


DIETS = {
    'mu_western_diet':    WESTERN_DIET,
    'mu_high_fiber_diet': HIGH_FIBER_DIET,
}


# ============================================================
# Helpers (gapseq SBML quirks)
# ============================================================
def find_exchange(model, cpd_id):
    """gapseq exchange naming: EX_cpd00027_e0 (sometimes EX_cpd00027_e)."""
    for suffix in ['_e0', '_e', '(e)', '_e_']:
        candidate = f'EX_{cpd_id}{suffix}'
        if candidate in model.reactions:
            return model.reactions.get_by_id(candidate)
    # fuzzy fallback: any EX reaction whose ID contains the cpd
    for rxn in model.exchanges:
        if cpd_id in rxn.id:
            return rxn
    return None


def find_biomass(model):
    """gapseq biomass reaction: usually bio1 or contains 'biomass' / cpd11416."""
    for hint in ['bio1', 'BIOMASS', 'biomass', 'cpd11416']:
        for rxn in model.reactions:
            if hint.lower() in rxn.id.lower():
                return rxn
    return None


def apply_diet(model, diet):
    """Close all exchanges, then open uptakes for diet metabolites."""
    opened, missing = 0, 0
    for rxn in model.exchanges:
        rxn.lower_bound = 0
    for cpd, uptake in diet.items():
        rxn = find_exchange(model, cpd)
        if rxn is None:
            missing += 1
            continue
        rxn.lower_bound = -float(uptake)
        opened += 1
    return opened, missing


# ============================================================
# Metadata
# ============================================================
def load_cc_map():
    wb = load_workbook(METADATA_XLSX, read_only=True)
    ws = wb['Cluster_assignments']
    rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    return dict(zip(df['sample'], df['CC']))


# ============================================================
# Main
# ============================================================
def main():
    cc_map = load_cc_map()
    gem_files = sorted(GEMS_DIR.glob('*.xml'))
    print(f"Found {len(gem_files)} GEM SBML files")
    print(f"Running FBA on {len(DIETS)} diet media: {', '.join(DIETS)}\n")

    rows = []
    for i, gem_path in enumerate(gem_files, 1):
        strain = gem_path.stem
        cc = cc_map.get(strain, 'unknown')
        row = {'strain': strain, 'CC': cc}
        print(f"[{i:3d}/{len(gem_files)}] {strain} ({cc})")

        try:
            model = cobra.io.read_sbml_model(str(gem_path))
        except Exception as e:
            print(f"  ! load failed: {e}")
            for col in DIETS:
                row[col] = np.nan
                row[f'status_{col}'] = 'load_error'
            rows.append(row)
            continue

        bio = find_biomass(model)
        if bio is None:
            print("  ! no biomass reaction found")
            for col in DIETS:
                row[col] = np.nan
                row[f'status_{col}'] = 'no_biomass'
            rows.append(row)
            continue

        model.objective = bio

        # FBA on each diet (model.copy() so diets do not leak across iterations)
        for col, diet in DIETS.items():
            try:
                m = model.copy()
                opened, missing = apply_diet(m, diet)
                sol = m.optimize()
                row[col] = float(sol.objective_value) if sol.status == 'optimal' else np.nan
                row[f'status_{col}']     = sol.status
                row[f'ex_opened_{col}']  = opened
                row[f'ex_missing_{col}'] = missing
                print(f"    {col}: mu = {row[col]:.4f}  ({sol.status})")
            except Exception as e:
                row[col] = np.nan
                row[f'status_{col}']     = f'error: {e}'
                row[f'ex_opened_{col}']  = 0
                row[f'ex_missing_{col}'] = 0
                print(f"    {col}: ERROR — {e}")

        rows.append(row)

    df = pd.DataFrame(rows)
    df.to_csv(OUTPUT_CSV, index=False)

    print(f"\n✓ Saved {OUTPUT_CSV}")
    print(f"  Strains: {len(df)}")
    for col in DIETS:
        valid = df[col].dropna()
        if len(valid):
            print(f"  {col}: {len(valid)} valid, mu range "
                  f"{valid.min():.4f} – {valid.max():.4f}")
        else:
            print(f"  {col}: no valid results")


if __name__ == '__main__':
    main()
