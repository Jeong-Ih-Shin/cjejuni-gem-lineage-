"""
Step 03 — Figure 1: GEM reaction-inventory lineage signal.

Inputs:
    intermediate/reaction_presence_absence.csv   (from step 01)
    data/metadata.xlsx     (Cluster_assignments sheet)

Outputs:
    output/figures/Figure1_gem_lineage.{png,svg}

Analyses:
    A — Pairwise Jaccard distance: same-CC vs different-CC distributions
        (one-sided Mann-Whitney U test)
    B — PCA of reaction-inventory binary vectors, colored by CC, bounding ellipses
    C — Per-CC mean within-CC vs between-CC Jaccard distance
"""
from pathlib import Path
import sys
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.spatial.distance import pdist, squareform, cdist
from scipy.stats import mannwhitneyu
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import (DATA_DIR, INTERMEDIATE_DIR, NPG, SEM, CC_COLORS,
                    apply_style, save_fig, bounding_ellipse)

apply_style()


def cc_label(cc_full):
    """Data string 'ST-21 complex' -> display label 'CC-21'."""
    return cc_full.replace(" complex", "").replace("ST-", "CC-")


def main():
    pa = pd.read_csv(INTERMEDIATE_DIR / 'reaction_presence_absence.csv', index_col=0).T
    pa = pa[~pa.index.str.contains('KCTC|ATCC', case=False, na=False)]

    wb = load_workbook(DATA_DIR / 'metadata.xlsx', read_only=True)
    ws = wb['Cluster_assignments']
    rows = list(ws.iter_rows(values_only=True))
    v4 = pd.DataFrame(rows[1:], columns=rows[0]).dropna(subset=['sample'])
    v4 = v4[~v4['sample'].str.contains('KCTC|ATCC', case=False, na=False)]
    cc_map = dict(zip(v4['sample'], v4['CC']))

    common = pa.index.intersection(set(cc_map.keys()))
    pa = pa.loc[common]
    cc = pd.Series([cc_map[s] for s in pa.index], index=pa.index, name='CC')
    print(f'Strains used: {len(pa)} | Reactions: {pa.shape[1]}')

    # Pairwise Jaccard
    samples = pa.index.tolist()
    dist_arr = pdist(pa.values.astype(int), metric='jaccard')
    dist_mat = pd.DataFrame(squareform(dist_arr), index=samples, columns=samples)
    pairs = []
    for i, s1 in enumerate(samples):
        for j, s2 in enumerate(samples):
            if i >= j:
                continue
            c1, c2 = cc[s1], cc[s2]
            if pd.isna(c1) or pd.isna(c2):
                continue
            pairs.append({'cc1': c1, 'cc2': c2, 'same_cc': c1 == c2,
                          'dist': dist_mat.iloc[i, j]})
    pairs_df = pd.DataFrame(pairs)
    same = pairs_df[pairs_df.same_cc]['dist'].values
    diff = pairs_df[~pairs_df.same_cc]['dist'].values
    _, p = mannwhitneyu(same, diff, alternative='less')
    med_same, med_diff = np.median(same), np.median(diff)
    print(f'Same-CC pairs: n={len(same)}, median={med_same:.3f}')
    print(f'Diff-CC pairs: n={len(diff)}, median={med_diff:.3f}')
    print(f'Mann-Whitney p = {p:.2e}')

    # PCA
    pca = PCA(n_components=2)
    coords = pca.fit_transform(pa.values.astype(float))
    var = pca.explained_variance_ratio_

    cc_counts = cc.value_counts()
    top_ccs = [c for c in cc_counts.head(8).index if c is not None and not pd.isna(c)]
    NO_ELLIPSE = {'ST-658 complex'}  # documented split structure (data string, keep as-is)

    # PANEL SWAP: A = per-CC validation (left), B = PCA (right)
    # width_ratios flipped so the wider PCA panel is now on the right.
    fig, axes = plt.subplots(1, 2, figsize=(6.69, 3.5), gridspec_kw={'width_ratios': [1, 1.5]})
    plt.subplots_adjust(top=0.88, bottom=0.13, left=0.08, right=0.98, wspace=0.5)

    # A — per-CC bars (validation)  [now axes[0], left]
    ax = axes[0]
    per_cc = []
    for this_cc in cc_counts[cc_counts >= 2].index:
        if pd.isna(this_cc) or this_cc is None:
            continue
        in_idx = cc[cc == this_cc].index
        out_idx = cc[(cc != this_cc) & cc.notna()].index
        in_mat = pa.loc[in_idx].values.astype(int)
        out_mat = pa.loc[out_idx].values.astype(int)
        if len(in_mat) < 2:
            continue
        within = pdist(in_mat, metric='jaccard').mean()
        between = cdist(in_mat, out_mat, metric='jaccard').mean()
        per_cc.append({'cc': f'{cc_label(this_cc)} (n={len(in_idx)})',
                       'within': within, 'between': between, 'n': len(in_idx)})
    per_cc_df = pd.DataFrame(per_cc).sort_values('n', ascending=True)
    y_pos = np.arange(len(per_cc_df))
    ax.barh(y_pos - 0.20, per_cc_df['within'], 0.38, color=NPG['red'],
            label='Within-CC mean', edgecolor='white', linewidth=1.2)
    ax.barh(y_pos + 0.20, per_cc_df['between'], 0.38, color=SEM['neutral'],
            label='Between-CC mean', edgecolor='white', linewidth=1.2)
    ax.set_yticks(y_pos)
    ax.set_yticklabels(per_cc_df['cc'])
    ax.set_xlabel('Mean Jaccard distance')
    ax.set_title('A', loc='left', x=-0.30, fontsize=11, fontweight='bold')
    ax.legend(frameon=False, loc='lower center', bbox_to_anchor=(0.5, 1.02),
              ncol=2, fontsize=7)
    ax.grid(axis='x', alpha=0.5, color=SEM['grid'], linewidth=0.5)
    ax.set_xlim(0, max(per_cc_df['between'].max(), per_cc_df['within'].max()) * 1.10)

    # B — PCA  [now axes[1], right]
    ax = axes[1]
    ax.grid(alpha=0.4, color=SEM['grid'], linewidth=0.5)
    for i, this_cc in enumerate(top_ccs):
        mask = (cc == this_cc).values
        sub = coords[mask]
        n = cc_counts[this_cc]
        if this_cc not in NO_ELLIPSE and n >= 2:
            bounding_ellipse(sub, ax, color=CC_COLORS[i])
        ax.scatter(sub[:, 0], sub[:, 1], c=[CC_COLORS[i]],
                   label=f'{cc_label(this_cc)} (n={n})',
                   s=40, alpha=0.85, edgecolor='white', linewidth=0.5, zorder=3)
    other_mask = (~cc.isin(top_ccs) & cc.notna()).values
    ax.scatter(coords[other_mask][:, 0], coords[other_mask][:, 1], c=SEM['neutral'],
               label=f'Other (n={other_mask.sum()})', s=25, alpha=0.6,
               edgecolor='white', linewidth=1, zorder=2)
    ax.set_xlabel(f'PC1 ({var[0]*100:.1f}%)')
    ax.set_ylabel(f'PC2 ({var[1]*100:.1f}%)')
    ax.set_title('B', loc='left', x=-0.10, fontsize=11, fontweight='bold')
    ax.legend(loc='upper left', bbox_to_anchor=(0.7, 1.05), fontsize=7, frameon=False)

    save_fig(fig, 'Figure1_gem_lineage')
    plt.close(fig)
    print('Done.')


if __name__ == '__main__':
    main()
