"""
Step 02 — Per-CC Fisher exact enrichment of GEM reactions.

For each target Clonal Complex (CC), tests every reaction in the pan-reactome:
    2×2 contingency:  (reaction present/absent) × (strain ∈ CC / ∉ CC)
    Fisher exact two-sided p-value.

Inputs:
    intermediate/reaction_presence_absence.csv   (from step 01)
    intermediate/reaction_names.csv              (from step 01)
    data/metadata.xlsx     (Cluster_assignments sheet)

Outputs:
    intermediate/{CC_NAME}_vs_others_enrichment.csv
        Columns: reaction, name, pct_in_cc, pct_others, pct_diff, p
        One file per target CC (default: ST-443, ST-658).
"""
from pathlib import Path
import sys
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import fisher_exact

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import DATA_DIR, INTERMEDIATE_DIR

# CCs analysed in the manuscript (Figure 3)
TARGET_CCS = ['ST-443 complex', 'ST-658 complex']


def load_inputs():
    pa = pd.read_csv(INTERMEDIATE_DIR / 'reaction_presence_absence.csv', index_col=0)
    name_df = pd.read_csv(INTERMEDIATE_DIR / 'reaction_names.csv')
    name_map = dict(zip(name_df['reaction'], name_df['name']))

    wb = load_workbook(DATA_DIR / 'metadata.xlsx', read_only=True)
    ws = wb['Cluster_assignments']
    rows = list(ws.iter_rows(values_only=True))
    v4 = pd.DataFrame(rows[1:], columns=rows[0]).dropna(subset=['sample'])
    cc_map = dict(zip(v4['sample'], v4['CC']))
    return pa, cc_map, name_map


def enrichment(pa, cc_map, target_cc, name_map):
    cohort = [s for s in pa.columns if s in cc_map and not pd.isna(cc_map[s])]
    in_cc  = [s for s in cohort if cc_map[s] == target_cc]
    out_cc = [s for s in cohort if cc_map[s] != target_cc]
    if len(in_cc) < 2 or len(out_cc) < 2:
        print(f'  Skipped {target_cc}: in_CC={len(in_cc)}, out_CC={len(out_cc)}')
        return

    short = target_cc.split()[0].replace('ST-', 'ST')   # e.g. ST443
    pct_in_col = f'pct_{short.lower()}'                 # pct_st443

    rows = []
    for rxn in pa.index:
        a = int(pa.loc[rxn, in_cc].sum())
        b = len(in_cc) - a
        c = int(pa.loc[rxn, out_cc].sum())
        d = len(out_cc) - c
        pct_in = 100 * a / len(in_cc)
        pct_out = 100 * c / len(out_cc)
        _, p = fisher_exact([[a, b], [c, d]], alternative='two-sided')
        rows.append({
            'reaction':  rxn,
            'name':      name_map.get(rxn, rxn),
            pct_in_col:  round(pct_in, 2),
            'pct_others': round(pct_out, 2),
            'pct_diff':  round(pct_in - pct_out, 2),
            'p':         p,
        })
    out = pd.DataFrame(rows).sort_values('p')
    out_path = INTERMEDIATE_DIR / f'{short}_vs_others_enrichment.csv'
    out.to_csv(out_path, index=False)
    n_sig = ((out['pct_diff'].abs() > 50) & (out['p'] < 0.01)).sum()
    print(f'  {target_cc}: {len(in_cc)} vs {len(out_cc)} strains, '
          f'{len(out)} reactions tested, {n_sig} significantly differential '
          f'(|pct_diff|>50, p<0.01)')
    print(f'    Saved: {out_path.relative_to(INTERMEDIATE_DIR.parent)}')


def main():
    pa, cc_map, name_map = load_inputs()
    print(f'Reactions × Strains: {pa.shape}')
    print(f'Strain IDs in matrix (sample): {list(pa.columns)[:5]} …')
    n_mapped = sum(1 for s in pa.columns if s in cc_map)
    print(f'Strains with CC assignment: {n_mapped}/{pa.shape[1]}')
    if n_mapped < pa.shape[1]:
        unmatched = [s for s in pa.columns if s not in cc_map]
        print(f'  [warn] No CC for: {unmatched[:5]}{"…" if len(unmatched)>5 else ""}')
        print(f'  Check that data/strain_id_map.xlsx covers all strains.')
    for cc in TARGET_CCS:
        enrichment(pa, cc_map, cc, name_map)


if __name__ == '__main__':
    main()
